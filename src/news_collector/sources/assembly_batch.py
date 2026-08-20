"""국회 회의록 zip(중첩 포함) 일괄 처리 → 선거별 이슈 salience + (후보×이슈) 발언.

2.5GB+ 회의록을 메모리 폭증 없이 처리한다: 모든 xlsx를 스트리밍으로 읽되, **선거
시점 창(window) 안의 발언만** 골라 카운트만 누적(발언 텍스트는 버림).

대수 번호로 매핑하지 않는다 — 발언 **날짜**가 어느 선거 창에 드는지로 슬라이스한다
(대선 번호 ≠ 국회 대수라 날짜 기준이 정확하고 누수도 안전).
"""

from __future__ import annotations

import io
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

from election_forecast.features.issue_matcher import IssueContextRule
from election_forecast.features.issue_matcher import match_issue_weights
from news_collector.sources.assembly_records import _week, clean_speaker, parse_meeting_date
from news_collector.sources.salience_base import normalize_within_election

_SPEECH_COLS = [f"발언내용{i}" for i in range(1, 8)]

# 6개 대선 창: (선거일 기준 ~D-90 ~ 선거 전날). 날짜로 슬라이스 = 누수 차단.
ELECTION_WINDOWS: dict[str, tuple[str, str]] = {
    "pres_2002": ("2002-09-20", "2002-12-18"),  # 16대 대선 2002-12-19
    "pres_2007": ("2007-09-20", "2007-12-18"),  # 17대 대선 2007-12-19
    "pres_2012": ("2012-09-20", "2012-12-18"),  # 18대 대선 2012-12-19
    "pres_2017": ("2017-02-08", "2017-05-08"),  # 19대 대선 2017-05-09
    "pres_2022": ("2021-12-09", "2022-03-08"),  # 20대 대선 2022-03-09
    "pres_2025": ("2025-03-05", "2025-06-02"),  # 21대 대선 2025-06-03
}


def which_election(d, windows: dict[str, tuple[str, str]]) -> str | None:
    """발언 날짜가 드는 선거 창의 election_id (없으면 None)."""

    iso = d.isoformat()
    for eid, (start, end) in windows.items():
        if start <= iso <= end:
            return eid
    return None


def iter_xlsx_bytes(zip_path: str | Path) -> Iterable[tuple[str, bytes]]:
    """마스터 zip의 모든 xlsx를 (중첩 zip 내부까지) (이름, bytes)로 산출."""

    with zipfile.ZipFile(zip_path) as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            if info.filename.lower().endswith(".xlsx"):
                yield info.filename, z.read(info)
            elif info.filename.lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(z.read(info))) as inner:
                    for j in inner.infolist():
                        if j.filename.lower().endswith(".xlsx"):
                            yield j.filename, inner.read(j)


def iter_rows_from_xlsx(data: bytes) -> Iterable[dict[str, Any]]:
    """xlsx bytes → 헤더-매핑된 dict 행 (read_only 스트리밍, 두 포맷 모두).

    - 16~20대: 단일 시트에 발언이 평면 저장.
    - 21·22대: '회의록목록' + 회의별 '{회의번호}_발언내용' 다중 시트. 발언내용 시트가
      회의일자·발언자·발언내용을 모두 보유하므로 그 시트들만 사용한다.

    read_only는 깨진 dimension(1×1)으로 행을 잘라버리므로 ``reset_dimensions()``로
    강제 전체 스캔한다. 헤더 행은 '회의일자'+('발언내용1' 또는 '발언자')로 식별한다.
    """

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.reset_dimensions()
            it = ws.iter_rows(values_only=True)
            header = None
            for _ in range(8):
                row = next(it, None)
                if row is None:
                    break
                cells = [str(c).strip() if c is not None else "" for c in row]
                if "회의일자" in cells and ("발언내용1" in cells or "발언자" in cells):
                    header = list(row)
                    break
            if header is None:
                continue  # 발언 없는 시트(회의록목록 등) 스킵
            for row in it:
                yield dict(zip(header, row))
    finally:
        wb.close()


def accumulate(
    rows: Iterable[dict[str, Any]],
    keyword_map: dict[str, list[str]],
    windows: dict[str, tuple[str, str]],
    member_to_slot: dict[str, str] | None,
    sal_counts: Counter,
    mem_counts: Counter,
    term_weights_by_election: dict[str, dict[tuple[str, str], float]] | None = None,
    issue_boosts_by_election: dict[str, dict[str, float]] | None = None,
    context_rules_by_election: dict[str, list[IssueContextRule]] | None = None,
) -> None:
    """행들을 선거 창으로 필터링하며 카운트만 누적 (텍스트 미보존)."""

    for r in rows:
        d = parse_meeting_date(r.get("회의일자"))
        if d is None:
            continue
        eid = which_election(d, windows)
        if eid is None:
            continue
        text = " ".join(str(r.get(c) or "") for c in _SPEECH_COLS if r.get(c))
        term_weights = (term_weights_by_election or {}).get(eid)
        issue_boosts = (issue_boosts_by_election or {}).get(eid)
        context_rules = (context_rules_by_election or {}).get(eid)
        issue_weights = match_issue_weights(
            text,
            keyword_map,
            term_weights=term_weights,
            issue_boosts=issue_boosts,
            context_rules=context_rules,
        )
        if not issue_weights:
            continue
        wk = _week(d)
        for iss, weight in issue_weights.items():
            sal_counts[(eid, iss, wk)] += weight
        if member_to_slot:
            spk = str(r.get("발언자") or "")
            name = clean_speaker(spk)
            last = name.split()[-1] if name.split() else ""  # 위원회 약칭 등 접두 제거 → 이름
            slot = (
                member_to_slot.get(name)
                or member_to_slot.get(last)
                or member_to_slot.get(str(r.get("의원ID") or ""))
            )
            if slot:
                for iss, weight in issue_weights.items():
                    mem_counts[(eid, slot, iss, wk)] += weight


def build_assembly_salience(
    zip_path: str | Path,
    keyword_map: dict[str, list[str]],
    windows: dict[str, tuple[str, str]] | None = None,
    member_to_slot: dict[str, str] | None = None,
    term_weights_by_election: dict[str, dict[tuple[str, str], float]] | None = None,
    issue_boosts_by_election: dict[str, dict[str, float]] | None = None,
    context_rules_by_election: dict[str, list[IssueContextRule]] | None = None,
    progress: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """전체 zip 처리 → (salience[선거 내 정규화], member_issue) DataFrame 반환."""

    windows = windows or ELECTION_WINDOWS
    sal_counts: Counter = Counter()
    mem_counts: Counter = Counter()
    for idx, (name, data) in enumerate(iter_xlsx_bytes(zip_path), 1):
        try:
            accumulate(
                iter_rows_from_xlsx(data),
                keyword_map,
                windows,
                member_to_slot,
                sal_counts,
                mem_counts,
                term_weights_by_election=term_weights_by_election,
                issue_boosts_by_election=issue_boosts_by_election,
                context_rules_by_election=context_rules_by_election,
            )
        except Exception as exc:  # noqa: BLE001 - one bad file must not abort the batch
            if progress:
                print(f"  [skip] {name.split('/')[-1]}: {exc}")
        if progress:
            print(f"  [{idx}] {name.split('/')[-1][:50]} | salience keys={len(sal_counts)}")

    sal_frames = []
    for eid in windows:
        rows = [{"issue_name": i, "period": w, "count": c} for (e, i, w), c in sal_counts.items() if e == eid]
        if rows:
            sal_frames.append(normalize_within_election(pd.DataFrame(rows), "count", eid, "assembly_speech"))
    salience = pd.concat(sal_frames, ignore_index=True) if sal_frames else pd.DataFrame()

    member = pd.DataFrame(
        [{"election_id": e, "slot": s, "issue_name": i, "period": w, "mentions": c} for (e, s, i, w), c in mem_counts.items()]
    )
    return salience, member
