"""Extract 15th National Assembly xlsx/PDF sources into JSON for xlsx export."""

from __future__ import annotations

import io
import json
import re
import zipfile
from pathlib import Path
from typing import Iterable

import openpyxl
from pypdf import PdfReader


DOWNLOADS = Path.home() / "Downloads"
OUTPUT_DIR = Path("outputs/15th_assembly_conversion")
OUTPUT_JSON = OUTPUT_DIR / "15th_assembly_extracted.json"
MAX_CELL_TEXT = 4800

BASE_COLUMNS = [
    "회의번호",
    "회의록구분",
    "대수",
    "회의구분",
    "위원회",
    "회수",
    "차수",
    "기타 정보",
    "회의일자",
    "안건",
    "발언자",
    "의원ID",
    "발언순번",
    "발언내용1",
    "발언내용2",
    "발언내용3",
    "발언내용4",
    "발언내용5",
    "발언내용6",
    "발언내용7",
]

PDF_EXTRA_COLUMNS = [
    "source_file",
    "source_page_start",
    "source_page_end",
    "extraction_method",
    "extraction_status",
    "text_quality",
    "notes",
]


def _compact_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _split_content(text: str) -> list[str]:
    text = _compact_text(text)
    parts = [text[i : i + MAX_CELL_TEXT] for i in range(0, len(text), MAX_CELL_TEXT)]
    parts = parts[:7]
    while len(parts) < 7:
        parts.append("")
    return parts


def _first_download_match(pattern: str) -> Path:
    matches = sorted(DOWNLOADS.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"no Downloads file matched {pattern!r}")
    return matches[0]


def read_plenary_xlsx() -> list[list[object]]:
    """Read the already-tabular 15th Assembly plenary workbook."""

    path = _first_download_match("*15*본회의*데이터셋.xlsx")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header = [str(value) if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
        index = {name: idx for idx, name in enumerate(header)}
        rows: list[list[object]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            out = []
            for column in BASE_COLUMNS:
                value = row[index[column]] if column in index and index[column] < len(row) else ""
                out.append(_compact_text(value))
            rows.append(out)
        return rows
    finally:
        workbook.close()


def _parse_filename_metadata(filename: str) -> dict[str, str]:
    date_match = re.search(r"\((\d{4})\.(\d{2})\.(\d{2})\.\)", filename)
    date = "-".join(date_match.groups()) if date_match else ""
    session_match = re.search(r"제(\d+)회\(([^)]+)\)", filename)
    session = f"제{session_match.group(1)}회" if session_match else ""
    meeting_type = session_match.group(2) if session_match else ("국정감사" if "국정감사" in filename else "")
    round_match = re.search(r"제(\d+)차\s*([^(]+?)\(전체회의\)", filename)
    sequence = f"제{round_match.group(1)}차" if round_match else ""
    committee = round_match.group(2).strip() if round_match else ""
    if not committee:
        audit_match = re.search(r"국정감사\s*([^(]+?)\(전체회의\)", filename)
        committee = audit_match.group(1).strip() if audit_match else ""
    return {
        "회의번호": "",
        "회의록구분": "국회",
        "대수": "15",
        "회의구분": meeting_type,
        "위원회": committee,
        "회수": session,
        "차수": sequence,
        "기타 정보": "PDF 추출",
        "회의일자": date,
        "안건": "",
    }


def _text_quality(text: str) -> tuple[str, float]:
    if not text:
        return "empty", 0.0
    hangul = len(re.findall(r"[가-힣]", text))
    bad = sum(1 for char in text if ord(char) in {0xA81A} or ord(char) < 32 and char not in "\n\t")
    ratio = hangul / max(len(text), 1)
    if len(text) >= 100 and ratio < 0.02:
        return "garbled", ratio
    if bad > 50:
        return "garbled", ratio
    return "ok", ratio


SPEAKER_RE = re.compile(
    r"(?m)^[\s\u3000]*(?:[◯○ㅇoO]\s*)"
    r"(?P<speaker>[가-힣A-Za-z0-9·ㆍ\-\s]{2,40}?)\s+"
    r"(?P<body>.+?)(?=^[\s\u3000]*(?:[◯○ㅇoO]\s*)[가-힣A-Za-z0-9·ㆍ\-\s]{2,40}?\s+|\Z)",
    re.DOTALL,
)


def _speaker_segments(text: str) -> list[tuple[str, str]]:
    segments: list[tuple[str, str]] = []
    for match in SPEAKER_RE.finditer(text):
        speaker = _compact_text(match.group("speaker"))
        body = _compact_text(match.group("body"))
        if len(body) >= 10:
            segments.append((speaker, body))
    return segments


def _iter_pdf_files_from_zip(zip_path: Path) -> Iterable[tuple[str, bytes]]:
    with zipfile.ZipFile(zip_path) as archive:
        for info in archive.infolist():
            if info.is_dir() or not info.filename.lower().endswith(".pdf"):
                continue
            yield info.filename, archive.read(info)


def extract_pdf_zip() -> tuple[list[list[object]], list[list[object]]]:
    """Extract all PDF files in the user-provided 15th Assembly zip."""

    zip_path = _first_download_match("*15*1996*행정위원회*.zip")
    rows: list[list[object]] = []
    summary: list[list[object]] = []
    for file_index, (filename, data) in enumerate(_iter_pdf_files_from_zip(zip_path), start=1):
        print(f"[pdf {file_index}] {filename}", flush=True)
        metadata = _parse_filename_metadata(filename)
        document_row_count = 0
        status_counts: dict[str, int] = {}
        try:
            reader = PdfReader(io.BytesIO(data))
            page_count = len(reader.pages)
            page_texts = []
            for page_index, page in enumerate(reader.pages, start=1):
                text = _compact_text(page.extract_text() or "")
                quality, ratio = _text_quality(text)
                status_counts[quality] = status_counts.get(quality, 0) + 1
                page_texts.append((page_index, text, quality, ratio))
        except Exception as exc:  # noqa: BLE001 - record extraction failures in workbook
            page_count = 0
            page_texts = [(0, "", "failed", 0.0)]
            status_counts["failed"] = 1
            metadata["기타 정보"] = f"PDF 추출 실패: {type(exc).__name__}: {exc}"

        for page_index, text, quality, ratio in page_texts:
            segments = _speaker_segments(text) if quality == "ok" else []
            if not segments:
                segments = [("", text)]
            for speaker, body in segments:
                base = [
                    metadata["회의번호"],
                    metadata["회의록구분"],
                    metadata["대수"],
                    metadata["회의구분"],
                    metadata["위원회"],
                    metadata["회수"],
                    metadata["차수"],
                    metadata["기타 정보"],
                    metadata["회의일자"],
                    metadata["안건"],
                    speaker,
                    "",
                    document_row_count + 1,
                    *_split_content(body),
                ]
                note = "" if quality == "ok" else "텍스트 추출 품질 낮음: OCR 또는 수동 확인 필요"
                rows.append(
                    [
                        *base,
                        filename,
                        page_index,
                        page_index,
                        "pypdf",
                        quality,
                        round(ratio, 4),
                        note,
                    ]
                )
                document_row_count += 1

        summary.append(
            [
                filename,
                page_count,
                document_row_count,
                status_counts.get("ok", 0),
                status_counts.get("garbled", 0),
                status_counts.get("empty", 0),
                status_counts.get("failed", 0),
            ]
        )
    return rows, summary


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    plenary_rows = read_plenary_xlsx()
    pdf_rows, pdf_summary = extract_pdf_zip()
    payload = {
        "plenary_columns": BASE_COLUMNS,
        "plenary_rows": plenary_rows,
        "pdf_columns": [*BASE_COLUMNS, *PDF_EXTRA_COLUMNS],
        "pdf_rows": pdf_rows,
        "summary_columns": [
            "source_file",
            "page_count",
            "extracted_rows",
            "ok_pages",
            "garbled_pages",
            "empty_pages",
            "failed_pages",
        ],
        "summary_rows": pdf_summary,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"saved {OUTPUT_JSON}")
    print(f"plenary_rows={len(plenary_rows)} pdf_rows={len(pdf_rows)} pdf_files={len(pdf_summary)}")


if __name__ == "__main__":
    main()
